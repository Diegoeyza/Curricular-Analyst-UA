import openpyxl
import re
import csv

# Load the workbook
file_path = r'C:\Users\diego\Downloads\Programación Maestro pruebas.xlsm'  # Replace with your Excel file path
workbook = openpyxl.load_workbook(file_path)

# Access the "MAESTRO" sheet
maestro_sheet = workbook['MAESTRO']

# Prepare to collect data
data_rows = []

# Find the first occurrence of a character in row 1 and save this column as START
start_col = None
for cell in maestro_sheet[1]:
    if cell.value and isinstance(cell.value, str) and re.search(r'[a-zA-Z]', cell.value):
        start_col = cell.column
        break

# Iterate through each row starting from row 3
for row in maestro_sheet.iter_rows(min_row=3, max_row=maestro_sheet.max_row):
    # Extract MATERIA/CURSO from column 4
    cell_value = str(row[3].value)
    if not cell_value or len(cell_value) < 7:
        continue

    # MATERIA/CURSO remains from column 4
    materia_curso = cell_value

    # Extract SECC from column N (14th column)
    secc = str(row[13].value)  # Column N is the 14th column (0-indexed as row[13])

    # Extract NRC from the next column (column 5)
    nrc = row[4].value

    # Extract TITULO from column 11 (column K)
    titulo = row[10].value

    # Iterate over the row to find data resembling "CL 13:30-15:20"
    for cell in row[start_col - 1:]:
        if cell.value and isinstance(cell.value, str) and re.search(r'\d{2}:\d{2}-\d{2}:\d{2}', cell.value):
            # Split the value into TIPO and HORA
            match = re.search(r'([A-Z\s]+)\s(\d{2}:\d{2}-\d{2}:\d{2})', cell.value)
            if match:
                tipo = match.group(1).strip()
                hora = match.group(2)

                # Extract FECHA from row 2 in the same column
                fecha = maestro_sheet.cell(row=2, column=cell.column).value

                # Collect the data row
                data_rows.append([nrc, materia_curso, secc, titulo, tipo, hora, fecha])

# Write the collected data to a CSV file
csv_file_path = 'output.csv'  # Adjust the path if needed
with open(csv_file_path, mode='w', newline='', encoding='utf-8') as file:
    writer = csv.writer(file)
    # Write header
    writer.writerow(['NRC', 'MATERIA/CURSO', 'SECC', 'TITULO', 'TIPO', 'HORA', 'FECHA'])
    # Write data rows
    writer.writerows(data_rows)

print(f"Data has been written to {csv_file_path}")
