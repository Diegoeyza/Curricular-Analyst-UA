import psycopg2
import pandas as pd
from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Database credentials
DB_CREDENTIALS = "dbname=links_ra user=postgres password=bduandes host=localhost port=5432"

# Establish connection to the database
def connect_db():
    try:
        connection = psycopg2.connect(DB_CREDENTIALS)
        return connection
    except Exception as e:
        messagebox.showerror("Database Connection Error", f"Error connecting to the database: {e}")
        return None

# The modifier can be WHERE or AND, and the prefix is the table prefix for filtering, like c. is the prefix of c.id
def user_input(data, base_query, modifier,prefix):
    if data:
        if "_" in data:
            # If the input is a number, filter by course ID (c.id)
            query = f"{base_query} {modifier} {prefix}id = '{data}'"
        else:
            # Otherwise, filter by course name (c.nombre)
            query = f"{base_query} {modifier} {prefix}nombre = '{data}'"
    else:
        query = base_query  # If no input, don't apply any filtering
    return query


# Function to execute queries
def execute_query(query):
    connection = connect_db()
    if not connection:
        return
    
    cursor = connection.cursor()
    try:
        cursor.execute(query)
        result = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]  # Get column names
        df = pd.DataFrame(result, columns=columns)  # Convert result to pandas DataFrame
        connection.commit()
        return df
    except Exception as e:
        messagebox.showerror("Query Execution Error", f"Error executing query: {e}")
        return None
    finally:
        cursor.close()
        connection.close()

# Function to display results in the UI
def show_results(df):
    if df is not None and not df.empty:
        result_text.delete(1.0, END)  # Clear previous results
        result_text.insert(END, df.to_string(index=False))  # Display the DataFrame as text
        global current_df
        current_df = df  # Store the DataFrame for later download
    else:
        messagebox.showinfo("No Results", "No results found or an error occurred.")

# Function to download the report as an xlsx file with Excel formatting
def download_report():
    if 'current_df' in globals() and current_df is not None:
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], initialdir="~/Documents")
        if file_path:
            # Export DataFrame to Excel
            current_df.to_excel(file_path, index=False, engine='openpyxl')

            # Open the Excel file and apply formatting
            wb = load_workbook(file_path)
            sheet = wb.active

            # Apply centering alignment to all cells
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # Apply bold formatting to the header row
            for cell in sheet[1]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = cell.font.copy(bold=True)

            # Save the changes
            wb.save(file_path)
            messagebox.showinfo("Download Complete", f"Report has been downloaded to {file_path}")
    else:
        messagebox.showwarning("No Data", "No data available to download. Please execute a query first.")

# Function for each predefined query
def query_1():
    course_input = course_input_var.get().strip()  # Get the input from the text box
    base_query = """
    SELECT nombre, id FROM courses
    """
    query=user_input(course_input,base_query, "WHERE", "")
    df = execute_query(query)
    show_results(df)

def query_2():
    course_input = course_input_var.get().strip()  # Get the input from the text box
    base_query = """
    SELECT nombre, id_objetivo, objetivo FROM objectives
    """
    query=user_input(course_input,base_query, "WHERE", "")
    df = execute_query(query)
    show_results(df)

def query_3():
    course_input = course_input_var.get().strip()  # Get the input from the text box
    base_query = """
    SELECT r.id, r.id_requisito FROM requirements r JOIN courses c ON c.id=r.id
    """
    query=user_input(course_input,base_query, "WHERE", "c.")
    df = execute_query(query)
    show_results(df)

# Function to filter query_4 by course name (c.nombre) or course id (c.id)
def query_4():
    course_input = course_input_var.get().strip()  # Get the input from the text box
    
    base_query = """
    SELECT 
        c.nombre AS curso,
        o.objetivo AS objetivo,
        ro.nombre AS prerequisito,
        oo.objetivo AS objetivo_prerequisito,
        rl.importancia
    FROM 
        ra_links rl
    JOIN 
        courses c ON rl.id = c.id
    JOIN 
        objectives o ON rl.id_objetivo = o.id_objetivo
    JOIN 
        courses ro ON rl.id_prerequisito = ro.id
    JOIN 
        objectives oo ON rl.id_objetivo_prerequisito = oo.id_objetivo
    """
    
    query=user_input(course_input,base_query, "WHERE", "c.")

    df = execute_query(query)
    show_results(df)

# Function for counting unlinked objectives per course with user input
def query_5():
    course_input = course_input_var.get().strip()  # Get the input from the text box
    
    base_query = """
    SELECT 
        c.nombre AS course_name,
        COUNT(o.id_objetivo) AS cantidad_obj
    FROM 
        objectives o
    JOIN 
        courses c ON o.id = c.id
    LEFT JOIN 
        ra_links rl1 ON o.id_objetivo = rl1.id_objetivo
    LEFT JOIN 
        ra_links rl2 ON o.id_objetivo = rl2.id_objetivo_prerequisito
    WHERE 
        rl1.id_objetivo IS NULL 
        AND rl2.id_objetivo_prerequisito IS NULL
    """
    
    # Add filtering condition if the user has provided input for course name or ID
    query=user_input(course_input,base_query, "AND", "c.")

    query += " GROUP BY c.nombre ORDER BY cantidad_obj DESC;"

    df = execute_query(query)
    show_results(df)

# Function to filter for courses where the input course is a prerequisite or outgoing objective
def query_6():
    course_input = course_input_var.get().strip()  # Get the input from the text box
    
    base_query = """
    SELECT 
        c.nombre AS curso,
        o.objetivo AS objetivo,
        ro.nombre AS prerequisito,
        oo.objetivo AS objetivo_prerequisito,
        rl.importancia
    FROM 
        ra_links rl
    JOIN 
        courses c ON rl.id = c.id
    JOIN 
        objectives o ON rl.id_objetivo = o.id_objetivo
    JOIN 
        courses ro ON rl.id_prerequisito = ro.id
    JOIN 
        objectives oo ON rl.id_objetivo_prerequisito = oo.id_objetivo
    """
    
    # Add filtering condition if the user has provided input for course name or ID
    query=user_input(course_input,base_query, "WHERE", "c.")

    df = execute_query(query)
    show_results(df)

# number of incoming and outgoing links for each course
def query_7():
    course_input = course_input_var.get().strip()  # Get the input from the text box
    
    base_query = """
    SELECT 
        c.id AS course_id, 
        c.nombre AS course_name,
        COALESCE(incoming_links.count, 0) AS incoming_links,
        COALESCE(outgoing_links.count, 0) AS outgoing_links,
        COALESCE(outgoing_links.count, 0) + COALESCE(incoming_links.count, 0) AS total_links
    FROM 
        courses c
    LEFT JOIN (
        SELECT 
            id_prerequisito AS course_id, 
            COUNT(*) AS count
        FROM 
            ra_links
        GROUP BY 
            id_prerequisito
    ) incoming_links ON c.id = incoming_links.course_id
    LEFT JOIN (
        SELECT 
            id AS course_id, 
            COUNT(*) AS count
        FROM 
            ra_links
        GROUP BY 
            id
    ) outgoing_links ON c.id = outgoing_links.course_id
    """
    
    # Add filtering condition if the user has provided input for course name or ID
    query=user_input(course_input,base_query, "WHERE", "c.")
    query += "ORDER BY total_links DESC;"

    df = execute_query(query)
    show_results(df)

def query_8():
    course_input = course_input_var.get().strip()  # Get the input from the text box
    base_query = """
    SELECT 
        c.nombre AS course_name,
        c.id AS course_id,
        o.id_objetivo AS objective_id,
        o.objetivo AS objective_description
    FROM 
        objectives o
    JOIN 
        courses c ON o.id = c.id
    LEFT JOIN 
        ra_links rl1 ON o.id_objetivo = rl1.id_objetivo
    LEFT JOIN 
        ra_links rl2 ON o.id_objetivo = rl2.id_objetivo_prerequisito
    WHERE 
        rl1.id_objetivo IS NULL 
        AND rl2.id_objetivo_prerequisito IS NULL
    """

    # Add filtering condition if the user has provided input for course name or ID
    query = user_input(course_input, base_query, "AND", "c.")

    query += " ORDER BY c.nombre, o.id_objetivo ASC;"

    df = execute_query(query)
    show_results(df)


def create_ui():
    root = Tk()
    root.title("Database Query Interface")

    # Center the window on the screen
    window_width = 900
    window_height = 700
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

    # Set the geometry and position the window in the center of the screen
    root.geometry(f"{window_width}x{window_height}+{(screen_width - window_width) // 2}+{(screen_height - window_height) // 2}")
    
    # Allow resizing by the user (this is default behavior)
    root.resizable(True, True)

    # Create a label
    label = Label(root, text="Select a predefined query to execute:", font=("Arial", 16))
    label.grid(row=0, column=0, columnspan=2, pady=20, sticky='n')

    # Create buttons for each query with better styling
    button1 = Button(root, text="Courses Query", command=query_1, width=25, height=2, font=("Arial", 12), relief="solid", bg="#4CAF50", fg="white")
    button1.grid(row=1, column=0, padx=10, pady=10)

    button2 = Button(root, text="Objectives Query", command=query_2, width=25, height=2, font=("Arial", 12), relief="solid", bg="#4CAF50", fg="white")
    button2.grid(row=1, column=1, padx=10, pady=10)

    button3 = Button(root, text="Requirements Query", command=query_3, width=25, height=2, font=("Arial", 12), relief="solid", bg="#4CAF50", fg="white")
    button3.grid(row=2, column=0, padx=10, pady=10)

    button4 = Button(root, text="RA Links Query", command=query_4, width=25, height=2, font=("Arial", 12), relief="solid", bg="#4CAF50", fg="white")
    button4.grid(row=2, column=1, padx=10, pady=10)

    button5 = Button(root, text="Number of Unlinked Objectives", command=query_5, width=40, height=2, font=("Arial", 12), relief="solid", bg="#4CAF50", fg="white")
    button5.grid(row=3, column=0, padx=10, pady=10)

    button6 = Button(root, text="Incoming and Outgoing Objective Links", command=query_6, width=40, height=2, font=("Arial", 12), relief="solid", bg="#4CAF50", fg="white")
    button6.grid(row=3, column=1, padx=10, pady=10)

    button7 = Button(root, text="Critical Courses", command=query_7, width=40, height=2, font=("Arial", 12), relief="solid", bg="#4CAF50", fg="white")
    button7.grid(row=4, column=0, padx=10, pady=10)

    button8 = Button(root, text="Unlinked Objectives", command=query_8, width=40, height=2, font=("Arial", 12), relief="solid", bg="#4CAF50", fg="white")
    button8.grid(row=4, column=1, padx=10, pady=10)

    # Entry field for user input (course name or ID)
    global course_input_var
    course_input_var = StringVar()
    entry_label = Label(root, text="Enter Course Name or ID (optional):", font=("Arial", 12))
    entry_label.grid(row=5, column=0, pady=10, padx=10)
    course_input = Entry(root, textvariable=course_input_var, font=("Arial", 12), width=30)
    course_input.grid(row=5, column=1, pady=10, padx=10)

    # Download button
    download_button = Button(root, text="Download Report as .xlsx", command=download_report, width=30, height=2, font=("Arial", 12), relief="solid", bg="#008CBA", fg="white")
    download_button.grid(row=6, column=0, columnspan=2, pady=10)

    # Text widget to show query results with scroll bar
    result_frame = Frame(root)
    result_frame.grid(row=7, column=0, columnspan=2, padx=10, pady=10, sticky='nsew')

    # Scrollbar for results
    scrollbar = Scrollbar(result_frame)
    scrollbar.grid(row=0, column=1, sticky='ns')  # Placing the scrollbar on the right side of the result frame

    # Text widget inside the result frame with monospace font
    global result_text
    result_text = Text(result_frame, wrap=WORD, height=15, width=90, font=("Courier New", 10))
    result_text.grid(row=0, column=0, sticky='nsew')  # Make text widget expand

    # Configure scrollbar to work with the text widget
    scrollbar.config(command=result_text.yview)
    result_text.config(yscrollcommand=scrollbar.set)

    # Center text inside the Text widget by setting alignment to center
    result_text.tag_configure("center", justify='center')
    result_text.insert(INSERT, "Your results will appear here...\n")
    result_text.tag_add("center", "1.0", "end")  # Center all inserted text

    # Allow the result frame to expand and fill available space
    result_frame.grid_rowconfigure(0, weight=1)
    result_frame.grid_columnconfigure(0, weight=1)

    # This ensures the window will expand properly without cropping
    root.grid_rowconfigure(7, weight=1)  # This makes row 7 (where the results frame is) expand with window resizing
    root.grid_columnconfigure(0, weight=1)
    root.grid_columnconfigure(1, weight=1)

    root.mainloop()

# Run the UI
if __name__ == "__main__":
    current_df = None  # To store the current DataFrame for download
    create_ui()
